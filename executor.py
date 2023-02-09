import openpyxl
from datetime import datetime, timedelta


class Parser_xls:
    def __init__(self, path_file):
        self.path_file: str = path_file
        self.start_time = None
        self.stop_time = None
        self.book = openpyxl.open(self.path_file,)
        self.sheet = self.book.worksheets[0]
        self.markets: dict = self.separate_market()
        self.check_time = None

    def set_time(self, start_time='00:00', stop_time='00:00'):
        self.start_time = Parser_xls.__validate_time(start_time)
        self.stop_time = Parser_xls.__validate_time(stop_time)
        if self.stop_time < self.start_time:
            self.check_time = self.__data_selection_for_start_time_greate
        if self.stop_time > self.start_time:
            self.check_time = self.__data_selection_for_stop_time_greate
        if self.stop_time == self.start_time:
            self.check_time = self.__data_selection_for_start_equals_stop

        # if self.stop_time == datetime(1900, 1, 1, 0, 0, 0):
        #     self.stop_time += timedelta(days=1)
        # print(self.stop_time)

    @staticmethod
    def __validate_time(str_time: str):
        return datetime.strptime(str_time, '%H:%M')

    def __data_selection_for_start_time_greate(self, time):
        return time >= self.start_time or time < self.stop_time

    def __data_selection_for_start_equals_stop(self, time):
        stop_time = self.stop_time + timedelta(days=1)
        return self.start_time <= time < stop_time

    def __data_selection_for_stop_time_greate(self, time):
        return self.start_time <= time < self.stop_time

    def separate_market(self):
        markets = {'ALL': 0}
        previous_value = ''
        for count, row in enumerate(self.sheet.iter_rows(min_row=0, max_row=self.sheet.max_row)):
            # print(count)
            # print(previous_value)
            now_value = str(row[0].value)
            if ('Чек ' in now_value and ' от ' in now_value) and \
                    ('Чек 'not in previous_value and ' от ' not in previous_value):
                markets[previous_value] = count
            previous_value = now_value
        return markets

    def pr_market(self):
        for i in self.markets:
            print(i)

    def get_all_market(self):
        markets = list(self.markets.keys())
        return markets

    def start_parse(self, market):
        breaker = False if market == "ALL" else True
        count = 0
        result = 0
        for row in self.sheet.iter_rows(min_row=self.markets[market]+1, max_row=self.sheet.max_row):
            receipt_details = str(row[0].value)
            if 'Чек ' and ' от ' in receipt_details:
                if row[1].value:
                    value = float(str(row[1].value).replace(' ', '').replace(',', '.'))
                    tm = datetime.strptime(receipt_details.split()[-1], '%H:%M:%S')
                    if self.check_time(tm):
                        print(row[1].value)
                        result += value
                        count += 1
            elif breaker:
                break
        return round(result, 2), count

    def create_report(self, market):
        # print(market, type(market))
        # print(summ, type(summ))
        try:
            name_sheet = f'{self.start_time.hour}.{self.start_time.minute}-' \
                         f'{self.stop_time.hour}.{self.stop_time.minute}'
            new_sheet = self.book.create_sheet(name_sheet, len(self.book.sheetnames))
            # print(self.book.sheetnames)

            if market == "ALL":
                for market in self.get_all_market():
                    self.__filling_table(sheet=new_sheet, market=market)
            else:
                self.__filling_table(sheet=new_sheet, market=market)
            print(self.path_file)
            self.book.save(self.path_file)
        except PermissionError:
            print('--- Oшибка! Вы забыли закрыть excel файл ---')

    def __filling_table(self, sheet, market):
        result = self.start_parse(market)
        row = sheet.max_row + 1  # ищем пустую строку в файле
        sheet["A" + str(row)] = market
        sheet["B" + str(row)] = result[0]
        sheet["C" + str(row)] = result[1]